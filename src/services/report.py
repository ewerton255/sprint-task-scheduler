from datetime import datetime, timedelta, time, timezone
from typing import Dict, List, Optional
from pathlib import Path
import json
from loguru import logger
import markdown
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, LongTable
from reportlab.lib.enums import TA_LEFT, TA_CENTER
from reportlab.platypus.flowables import KeepTogether
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

from ..models.entities import Task, UserStory, Sprint, TaskStatus, WorkFront, SprintMetrics
from ..models.config import DayOff, ExecutorsConfig

class ReportGenerator:
    """Serviço responsável pela geração de relatórios"""

    def __init__(self, sprint: Sprint, dayoffs: Dict[str, List[DayOff]], output_dir: str, team_name: str, executors: ExecutorsConfig):
        """
        Inicializa o gerador de relatórios
        
        Args:
            sprint: Sprint a ser relatada
            dayoffs: Dicionário de ausências por executor
            output_dir: Diretório de saída dos relatórios
            team_name: Nome da equipe
            executors: Configuração dos executores
        """
        self.sprint = sprint
        self.dayoffs = dayoffs
        self.output_dir = Path(output_dir)
        self.team_name = team_name
        self.executors = executors
        self.metrics = sprint.metrics
        
        # Cria o diretório de saída se não existir
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Define os estilos do PDF
        self.styles = getSampleStyleSheet()
        self._setup_styles()
        
        # Define as cores para o Excel
        self.excel_colors = {
            'weekend': PatternFill(start_color='FFB3B3', end_color='FFB3B3', fill_type='solid'),  # Vermelho claro
            'dayoff': PatternFill(start_color='B3B3B3', end_color='B3B3B3', fill_type='solid'),   # Cinza claro
            'full': PatternFill(start_color='B3FFB3', end_color='B3FFB3', fill_type='solid'),     # Verde claro
            'partial': PatternFill(start_color='B3D1FF', end_color='B3D1FF', fill_type='solid'),  # Azul claro
            'empty': PatternFill(start_color='FFFFB3', end_color='FFFFB3', fill_type='solid')     # Amarelo claro
        }
        
        # Define os horários dos períodos
        self.morning_start = datetime.strptime('09:00', '%H:%M').time()
        self.morning_end = datetime.strptime('12:00', '%H:%M').time()
        self.afternoon_start = datetime.strptime('14:00', '%H:%M').time()
        self.afternoon_end = datetime.strptime('17:00', '%H:%M').time()

    def _setup_styles(self):
        """Configura estilos personalizados para o relatório"""
        # Estilo para o título principal
        self.styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=self.styles['Title'],
            fontSize=16,
            spaceAfter=30,
            textColor=colors.HexColor('#FF6B00'),  # Laranja
            alignment=TA_CENTER
        ))
        
        # Estilo para títulos de seção
        self.styles.add(ParagraphStyle(
            name='CustomHeading1',
            parent=self.styles['Heading1'],
            fontSize=14,
            spaceAfter=12,
            textColor=colors.HexColor('#FF6B00'),  # Laranja
            alignment=TA_LEFT
        ))
        
        # Estilo para subtítulos
        self.styles.add(ParagraphStyle(
            name='CustomHeading2',
            parent=self.styles['Heading2'],
            fontSize=12,
            spaceAfter=6,
            textColor=colors.HexColor('#FF8533'),  # Laranja mais claro
            alignment=TA_LEFT
        ))
        
        # Estilo para texto normal com quebra de linha
        self.styles.add(ParagraphStyle(
            name='NormalWrap',
            parent=self.styles['Normal'],
            fontSize=10,
            leading=12,
            spaceAfter=6,
            alignment=TA_LEFT
        ))
        
        # Estilo para células de tabela
        self.styles.add(ParagraphStyle(
            name='TableCell',
            parent=self.styles['Normal'],
            fontSize=9,
            leading=11,
            alignment=TA_LEFT
        ))
        
        # Estilo para cabeçalhos de tabela
        self.styles.add(ParagraphStyle(
            name='TableHeader',
            parent=self.styles['Normal'],
            fontSize=10,
            leading=12,
            alignment=TA_LEFT,
            fontName='Helvetica-Bold',
            textColor=colors.white
        ))

    def _create_table_style(self, header_bg_color=colors.HexColor('#FF6B00')):  # Laranja
        """Cria um estilo padrão para as tabelas"""
        return TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), header_bg_color),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('WORDWRAP', (0, 0), (-1, -1), True),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FFF5EB')]),  # Branco e laranja muito claro
        ])

    def _count_working_days(self, start_date: datetime, end_date: datetime) -> int:
        """
        Conta o número de dias úteis entre duas datas (excluindo finais de semana)
        
        Args:
            start_date: Data inicial
            end_date: Data final
            
        Returns:
            int: Número de dias úteis
        """
        working_days = 0
        current_date = start_date
        while current_date <= end_date:
            # 5 = Sábado, 6 = Domingo
            if current_date.weekday() < 5:
                working_days += 1
            current_date += timedelta(days=1)
        return working_days

    def _generate_markdown(self) -> str:
        """Gera o conteúdo do relatório em Markdown"""
        report = []
        
        # Título
        report.append(f"# Relatório de Agendamento - Sprint {self.sprint.name}")
        report.append("")
        
        # 1. Resumo Geral da Sprint
        report.append("## 1. Resumo Geral da Sprint")
        report.append("")
        report.append(f"- **Sprint:** {self.sprint.name}")
        report.append(f"- **Início:** {self.sprint.start_date.strftime('%d/%m/%Y')}")
        report.append(f"- **Término:** {self.sprint.end_date.strftime('%d/%m/%Y')}")
        report.append(f"- **Total de User Stories:** {len(self.sprint.user_stories)}")
        report.append("")
        
        # 2. User Stories Planejadas
        report.append("## 2. User Stories Planejadas")
        report.append("")
        report.append("| ID | Título | Responsável | Data de Finalização | Story Points |")
        report.append("|----|--------|-------------|---------------------|--------------|")
        
        for us in self.sprint.user_stories:
            end_date = us.end_date.strftime('%d/%m/%Y') if us.end_date else '-'
            report.append(
                f"| {us.id} | {us.title} | {us.assignee or '-'} | {end_date} | {us.story_points or '-'} |"
            )
        
        report.append("")
        
        # 3. Tasks não planejadas
        if self.metrics.not_scheduled_tasks:
            report.append("## 3. Tasks não planejadas")
            report.append("")
            report.append("| ID | Título | User Story | Motivo |")
            report.append("|----|--------|------------|--------|")
            
            for task in self.metrics.not_scheduled_tasks:
                report.append(
                    f"| {task['task_id']} | {task['title']} | {task['user_story_id']} | {task['reason']} |"
                )
            report.append("")
            
        # 4. Capacity dos Executores
        report.append("## 4. Capacity dos Executores")
        report.append("")
        report.append("| Executor | Capacity Total | Capacity Utilizada | Capacity Disponível | Datas de Ausência |")
        report.append("|----------|----------------|-------------------|---------------------|-------------------|")
        
        # Obtém todos os executores únicos de todas as frentes
        all_executors = set()
        for front in WorkFront:
            executors_list = getattr(self.executors, front.value, [])
            all_executors.update(executors_list)
        
        # Ordena os executores por email
        for executor in sorted(all_executors, key=lambda e: e.email):
            total = self.metrics.total_capacity.get(executor.email, 0)
            used = self.metrics.used_capacity.get(executor.email, 0)
            available = self.metrics.available_capacity.get(executor.email, 0)
            
            # Obtém as ausências do executor
            absences = []
            # Tenta encontrar as ausências ignorando case
            executor_dayoffs = next((dayoffs for name, dayoffs in self.dayoffs.items() 
                                  if name.lower() == executor.email.lower()), [])
            
            for dayoff in executor_dayoffs:
                period = {
                    "full": "dia inteiro",
                    "morning": "manhã",
                    "afternoon": "tarde"
                }
                absences.append(f"{dayoff.date.strftime('%d/%m/%Y')} ({period[dayoff.period]})")
            
            report.append(
                f"| {executor.email} | {total:.1f}h | {used:.1f}h | {available:.1f}h | {', '.join(absences) or '-'} |"
            )
        
        report.append("")
        
        # 5. Percentual de Capacity Preenchida
        report.append("## 5. Percentual de Capacity Preenchida")
        report.append("")
        report.append("| Métrica | Valor |")
        report.append("|---------|-------|")
        
        # Calcula o total de capacity disponível e utilizada
        total_available = sum(self.metrics.total_capacity.values())
        total_used = sum(self.metrics.used_capacity.values())
        
        # Calcula o percentual preenchido
        percent_filled = (total_used / total_available * 100) if total_available > 0 else 0
        
        report.append(f"| Percentual de Capacity Preenchida | {percent_filled:.2f}% |")
        report.append(f"| Total de Capacity Disponível | {total_available:.1f}h |")
        report.append(f"| Total de Capacity Utilizada | {total_used:.1f}h |")
        report.append("")
        
        return "\n".join(report)

    def generate(self) -> None:
        """Gera o relatório da sprint em PDF, Markdown e Excel"""
        # Gera o relatório em Markdown
        markdown_content = self._generate_markdown()
        markdown_path = self.output_dir / f"relatorio_sprint_{self.sprint.name.replace(' ', '_')}.md"
        markdown_path.write_text(markdown_content, encoding='utf-8')
        logger.info(f"Relatório Markdown gerado em {markdown_path}")
        
        # Cria o documento PDF
        pdf_path = self.output_dir / f"relatorio_sprint_{self.sprint.name.replace(' ', '_')}.pdf"
        doc = SimpleDocTemplate(
            str(pdf_path),
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )
        
        # Lista de elementos do documento
        elements = []
        
        # Título
        elements.append(Paragraph(f"Relatório da Sprint: {self.sprint.name} - {self.team_name}", self.styles['CustomTitle']))
        elements.append(Spacer(1, 12))
        
        # 1. Resumo Geral da Sprint
        elements.append(Paragraph("1. Resumo Geral da Sprint", self.styles['CustomHeading1']))
        elements.append(Paragraph(f"Sprint: {self.sprint.name}", self.styles['NormalWrap']))
        elements.append(Paragraph(
            f"Período: {self.sprint.start_date.strftime('%d/%m/%Y')} a {self.sprint.end_date.strftime('%d/%m/%Y')}",
            self.styles['NormalWrap']
        ))
        elements.append(Paragraph(f"Total de User Stories planejadas: {len(self.sprint.user_stories)}", self.styles['NormalWrap']))
        elements.append(Spacer(1, 12))
        
        # 2. User Stories Planejadas
        elements.append(Paragraph("2. User Stories Planejadas", self.styles['CustomHeading1']))
        us_data = [[
            Paragraph('ID', self.styles['TableHeader']),
            Paragraph('Título', self.styles['TableHeader']),
            Paragraph('Responsável', self.styles['TableHeader']),
            Paragraph('Data de Finalização', self.styles['TableHeader']),
            Paragraph('Story Points', self.styles['TableHeader'])
        ]]
        
        for us in self.sprint.user_stories:
            end_date = us.end_date.strftime('%d/%m/%Y') if us.end_date else '-'
            us_data.append([
                us.id,
                Paragraph(us.title, self.styles['TableCell']),
                Paragraph(us.assignee or '-', self.styles['TableCell']),
                end_date,
                str(us.story_points or '-')
            ])
        
        # Calcula larguras proporcionais
        available_width = doc.width
        us_table = LongTable(
            us_data,
            colWidths=[
                available_width * 0.1,  # ID
                available_width * 0.5,  # Título
                available_width * 0.15, # Responsável
                available_width * 0.15, # Data
                available_width * 0.1   # Story Points
            ]
        )
        us_table.setStyle(self._create_table_style())
        elements.append(KeepTogether(us_table))
        elements.append(Spacer(1, 12))
        
        # 3. Tasks não planejadas
        if self.metrics.not_scheduled_tasks:
            elements.append(Paragraph("3. Tasks não planejadas", self.styles['CustomHeading1']))
            
            not_scheduled_data = [[
                Paragraph('ID', self.styles['TableHeader']),
                Paragraph('Título', self.styles['TableHeader']),
                Paragraph('User Story', self.styles['TableHeader']),
                Paragraph('Motivo', self.styles['TableHeader'])
            ]]
            
            for task in self.metrics.not_scheduled_tasks:
                not_scheduled_data.append([
                    Paragraph(task['task_id'], self.styles['TableCell']),
                    Paragraph(task['title'], self.styles['TableCell']),
                    Paragraph(task['user_story_id'], self.styles['TableCell']),
                    Paragraph(task['reason'], self.styles['TableCell'])
                ])
            
            not_scheduled_table = LongTable(
                not_scheduled_data,
                colWidths=[
                    available_width * 0.1,  # ID
                    available_width * 0.4,  # Título
                    available_width * 0.2,  # User Story
                    available_width * 0.3   # Motivo
                ]
            )
            not_scheduled_table.setStyle(self._create_table_style())
            elements.append(KeepTogether(not_scheduled_table))
            elements.append(Spacer(1, 12))
            
        # 4. Capacity dos Executores
        elements.append(Paragraph("4. Capacity dos Executores", self.styles['CustomHeading1']))
        
        capacity_data = [[
            Paragraph('Executor', self.styles['TableHeader']),
            Paragraph('Capacity Total', self.styles['TableHeader']),
            Paragraph('Capacity Utilizada', self.styles['TableHeader']),
            Paragraph('Capacity Disponível', self.styles['TableHeader']),
            Paragraph('Datas de Ausência', self.styles['TableHeader'])
        ]]
        
        # Obtém todos os executores únicos de todas as frentes
        all_executors = set()
        for front in WorkFront:
            executors_list = getattr(self.executors, front.value, [])
            all_executors.update(executors_list)
        
        # Ordena os executores por email
        for executor in sorted(all_executors, key=lambda e: e.email):
            total = self.metrics.total_capacity.get(executor.email, 0)
            used = self.metrics.used_capacity.get(executor.email, 0)
            available = self.metrics.available_capacity.get(executor.email, 0)
            
            # Obtém as ausências do executor
            absences = []
            # Tenta encontrar as ausências ignorando case
            executor_dayoffs = next((dayoffs for name, dayoffs in self.dayoffs.items() 
                                  if name.lower() == executor.email.lower()), [])
            
            for dayoff in executor_dayoffs:
                period = {
                    "full": "dia inteiro",
                    "morning": "manhã",
                    "afternoon": "tarde"
                }
                absences.append(f"{dayoff.date.strftime('%d/%m/%Y')} ({period[dayoff.period]})")
            
            capacity_data.append([
                Paragraph(executor.email, self.styles['TableCell']),
                Paragraph(f"{total:.1f}h", self.styles['TableCell']),
                Paragraph(f"{used:.1f}h", self.styles['TableCell']),
                Paragraph(f"{available:.1f}h", self.styles['TableCell']),
                Paragraph(', '.join(absences) or '-', self.styles['TableCell'])
            ])
        
        capacity_table = LongTable(
            capacity_data,
            colWidths=[
                available_width * 0.25,  # Executor
                available_width * 0.15,  # Capacity Total
                available_width * 0.15,  # Capacity Utilizada
                available_width * 0.15,  # Capacity Disponível
                available_width * 0.3    # Datas de Ausência
            ]
        )
        capacity_table.setStyle(self._create_table_style())
        elements.append(KeepTogether(capacity_table))
        elements.append(Spacer(1, 12))
        
        # 5. Percentual de Capacity Preenchida
        elements.append(Paragraph("5. Percentual de Capacity Preenchida", self.styles['CustomHeading1']))
        
        # Calcula o total de capacity disponível e utilizada
        total_available = sum(self.metrics.total_capacity.values())
        total_used = sum(self.metrics.used_capacity.values())
        
        # Calcula o percentual preenchido
        percent_filled = (total_used / total_available * 100) if total_available > 0 else 0
        
        capacity_summary_data = [[
            Paragraph('Métrica', self.styles['TableHeader']),
            Paragraph('Valor', self.styles['TableHeader'])
        ]]
        
        capacity_summary_data.extend([
            [Paragraph('Percentual de Capacity Preenchida', self.styles['TableCell']),
             Paragraph(f"{percent_filled:.2f}%", self.styles['TableCell'])],
            [Paragraph('Total de Capacity Disponível', self.styles['TableCell']),
             Paragraph(f"{total_available:.1f}h", self.styles['TableCell'])],
            [Paragraph('Total de Capacity Utilizada', self.styles['TableCell']),
             Paragraph(f"{total_used:.1f}h", self.styles['TableCell'])]
        ])
        
        capacity_summary_table = LongTable(
            capacity_summary_data,
            colWidths=[
                available_width * 0.6,  # Métrica
                available_width * 0.4   # Valor
            ]
        )
        capacity_summary_table.setStyle(self._create_table_style())
        elements.append(KeepTogether(capacity_summary_table))
        elements.append(Spacer(1, 12))
        
        # Gera o PDF
        doc.build(elements)
        logger.info(f"Relatório PDF gerado em {pdf_path}")
        
        # Gera o relatório em Excel
        self._generate_excel()

    def _generate_excel(self) -> None:
        """Gera o relatório da sprint em Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Sprint {self.sprint.name}"

        sprint_start = self.sprint.start_date
        sprint_end = self.sprint.end_date
        if sprint_start.tzinfo is None:
            sprint_start = sprint_start.replace(tzinfo=timezone.utc)
        if sprint_end.tzinfo is None:
            sprint_end = sprint_end.replace(tzinfo=timezone.utc)

        all_executors = set()
        for front in WorkFront:
            executors_list = getattr(self.executors, front.value, [])
            all_executors.update(executors_list)
        sorted_executors = sorted(all_executors, key=lambda e: e.email)

        num_date_columns = (sprint_end - sprint_start).days + 1
        last_col = num_date_columns + 1
        last_col_letter = get_column_letter(last_col)

        # Bordas
        medium = Side(style='medium')
        thin = Side(style='thin')
        none = Side(style=None)

        # Largura fixa para todas as colunas (inclusive datas e label)
        col_width = 15
        for col in range(1, last_col + 1):
            ws.column_dimensions[get_column_letter(col)].width = col_width

        current_row = 1
        for executor in sorted_executors:
            bloco_inicio = current_row
            bloco_fim = current_row + 4  # email, vazia, datas, manhã, tarde

            # Linha 1: Email do executor (mesclado de A até última data)
            ws.merge_cells(f'A{current_row}:{last_col_letter}{current_row}')
            ws.cell(row=current_row, column=1, value=executor.email)
            ws.cell(row=current_row, column=1).font = Font(bold=True)
            ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center')
            current_row += 1

            # Linha 2: Vazia (mesclada de A até última data)
            ws.merge_cells(f'A{current_row}:{last_col_letter}{current_row}')
            current_row += 1

            # Linha 3: Datas (A vazia, datas de B em diante)
            ws.cell(row=current_row, column=1, value=None)
            ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center')
            current_date = sprint_start
            for col in range(2, last_col + 1):
                excel_date = current_date.replace(tzinfo=None)
                cell = ws.cell(row=current_row, column=col, value=excel_date)
                cell.number_format = 'dd/mm/yyyy'
                cell.alignment = Alignment(horizontal='center')
                current_date += timedelta(days=1)
            current_row += 1

            # Linha 4: Manhã
            ws.cell(row=current_row, column=1, value="Manhã")
            ws.cell(row=current_row, column=1).font = Font(bold=True)
            ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center')
            current_date = sprint_start
            for col in range(2, last_col + 1):
                is_weekend = current_date.weekday() >= 5
                executor_dayoffs = next((dayoffs for name, dayoffs in self.dayoffs.items() if name.lower() == executor.email.lower()), [])
                dayoff = next((d for d in executor_dayoffs if d.date.date() == current_date.date()), None)
                morning_allocation = self._calculate_period_allocation(
                    executor.email, current_date, self.morning_start, self.morning_end
                )
                cell = ws.cell(row=current_row, column=col)
                if is_weekend:
                    cell.fill = self.excel_colors['weekend']
                elif dayoff and dayoff.period in ['full', 'morning']:
                    cell.fill = self.excel_colors['dayoff']
                else:
                    self._apply_allocation_color(cell, morning_allocation)
                current_date += timedelta(days=1)
            current_row += 1

            # Linha 5: Tarde
            ws.cell(row=current_row, column=1, value="Tarde")
            ws.cell(row=current_row, column=1).font = Font(bold=True)
            ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center')
            current_date = sprint_start
            for col in range(2, last_col + 1):
                is_weekend = current_date.weekday() >= 5
                executor_dayoffs = next((dayoffs for name, dayoffs in self.dayoffs.items() if name.lower() == executor.email.lower()), [])
                dayoff = next((d for d in executor_dayoffs if d.date.date() == current_date.date()), None)
                afternoon_allocation = self._calculate_period_allocation(
                    executor.email, current_date, self.afternoon_start, self.afternoon_end
                )
                cell = ws.cell(row=current_row, column=col)
                if is_weekend:
                    cell.fill = self.excel_colors['weekend']
                elif dayoff and dayoff.period in ['full', 'afternoon']:
                    cell.fill = self.excel_colors['dayoff']
                else:
                    self._apply_allocation_color(cell, afternoon_allocation)
                current_date += timedelta(days=1)
            current_row += 1

            # Aplicar borda média externa ao bloco do executor (inclui email, vazia, datas, manhã, tarde)
            for row in range(bloco_inicio, bloco_fim + 1):
                for col in range(1, last_col + 1):
                    cell = ws.cell(row=row, column=col)
                    # Borda externa média somente nas extremidades do bloco
                    left = medium if col == 1 else none
                    right = medium if col == last_col else none
                    top = medium if row == bloco_inicio else none
                    bottom = medium if row == bloco_fim else none

                    # Células de e-mail e labels não recebem borda interna
                    if (row in (bloco_inicio, bloco_inicio + 1, bloco_inicio + 3, bloco_inicio + 4) and col == 1) or (row == bloco_inicio and col > 1):
                        cell.border = Border(left=left, right=right, top=top, bottom=bottom)
                    # Células de datas e períodos recebem borda interna fina
                    else:
                        cell.border = Border(
                            left=left if left != none else thin,
                            right=right if right != none else thin,
                            top=top if top != none else thin,
                            bottom=bottom if bottom != none else thin
                        )

            # --- REGRAS EXTRAS PARA GARANTIR O VISUAL ---
            # 1. Remover borda inferior da linha do e-mail (todas as células)
            for col in range(1, last_col + 1):
                cell = ws.cell(row=bloco_inicio, column=col)
                b = cell.border
                cell.border = Border(
                    left=b.left, right=b.right, top=b.top, bottom=none
                )
            # 2. Remover borda superior da linha vazia (todas as células)
            for col in range(1, last_col + 1):
                cell = ws.cell(row=bloco_inicio + 1, column=col)
                b = cell.border
                cell.border = Border(
                    left=b.left, right=b.right, top=none, bottom=b.bottom
                )
            # 3. Remover borda superior e inferior das células da coluna A entre o topo do bloco e a label 'Manhã'
            for row in (bloco_inicio + 1, bloco_inicio + 2):
                cell = ws.cell(row=row, column=1)
                b = cell.border
                cell.border = Border(
                    left=b.left, right=b.right, top=none, bottom=none
                )
            # 4. Remover borda superior da label 'Manhã'
            cell = ws.cell(row=bloco_inicio + 3, column=1)
            b = cell.border
            cell.border = Border(
                left=b.left, right=b.right, top=none, bottom=b.bottom
            )

            # Espaçamento entre executores
            current_row += 2

        # Legenda como bloco mesclado com borda média
        legend_row = current_row + 1
        legend_last_col = 4
        legend_last_col_letter = get_column_letter(legend_last_col)
        legend_items = [
            ("Fim de Semana", self.excel_colors['weekend']),
            ("Ausência", self.excel_colors['dayoff']),
            ("Alocação Completa (≥3h)", self.excel_colors['full']),
            ("Alocação Parcial (>0h)", self.excel_colors['partial']),
            ("Sem Alocação", self.excel_colors['empty'])
        ]
        ws.merge_cells(f'A{legend_row}:{legend_last_col_letter}{legend_row}')
        ws.cell(row=legend_row, column=1, value="Legenda:")
        ws.cell(row=legend_row, column=1).font = Font(bold=True)
        for i, (label, color) in enumerate(legend_items):
            row = legend_row + i + 1
            ws.merge_cells(f'A{row}:C{row}')
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=4).fill = color
        # Borda média ao redor do bloco da legenda
        for row in range(legend_row, legend_row + len(legend_items) + 1):
            for col in range(1, legend_last_col + 1):
                cell = ws.cell(row=row, column=col)
                left = medium if col == 1 else none
                right = medium if col == legend_last_col else none
                top = medium if row == legend_row else none
                bottom = medium if row == legend_row + len(legend_items) else none
                cell.border = Border(
                    left=left if left != none else thin,
                    right=right if right != none else thin,
                    top=top if top != none else thin,
                    bottom=bottom if bottom != none else thin
                )

        excel_path = self.output_dir / f"relatorio_sprint_{self.sprint.name.replace(' ', '_')}.xlsx"
        wb.save(str(excel_path))
        logger.info(f"Relatório Excel gerado em {excel_path}")

    def _calculate_period_allocation(self, executor_email: str, date: datetime, start_time: time, end_time: time) -> float:
        """
        Calcula a alocação de um executor em um período específico
        
        Args:
            executor_email: Email do executor
            date: Data a ser verificada
            start_time: Horário de início do período
            end_time: Horário de fim do período
            
        Returns:
            float: Horas alocadas no período
        """
        # Obtém todas as tasks do executor
        tasks = self.sprint.get_tasks_by_assignee(executor_email)
        
        # Filtra apenas tasks ativas
        active_tasks = [t for t in tasks if t.status not in [TaskStatus.CLOSED, TaskStatus.CANCELLED]]
        
        # Calcula horas alocadas no período
        allocated_hours = 0
        for task in active_tasks:
            if task.start_date and task.end_date:
                # Converte as datas para o mesmo timezone da data de referência
                task_start = task.start_date.astimezone(date.tzinfo)
                task_end = task.end_date.astimezone(date.tzinfo)
                
                # Verifica se a task está no dia
                if task_start.date() <= date.date() <= task_end.date():
                    # Calcula sobreposição com o período
                    period_start = datetime.combine(date.date(), start_time).replace(tzinfo=date.tzinfo)
                    period_end = datetime.combine(date.date(), end_time).replace(tzinfo=date.tzinfo)
                    
                    # Ajusta as datas para considerar apenas o período de trabalho
                    period_start = max(period_start, task_start)
                    period_end = min(period_end, task_end)
                    
                    if period_start < period_end:
                        # Calcula horas sobrepostas
                        overlap_hours = (period_end - period_start).total_seconds() / 3600
                        allocated_hours += overlap_hours
        
        return allocated_hours

    def _apply_allocation_color(self, cell, allocation: float) -> None:
        """
        Aplica a cor apropriada baseada na alocação
        
        Args:
            cell: Célula do Excel
            allocation: Horas alocadas
        """
        if allocation >= 3:
            cell.fill = self.excel_colors['full']
        elif allocation > 0:
            cell.fill = self.excel_colors['partial']
        else:
            cell.fill = self.excel_colors['empty'] 